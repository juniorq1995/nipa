import pandas as pd
import numpy as np
import os, math
from matplotlib import pyplot as plt
import time
from openpyxl import Workbook
from openpyxl.cell import get_column_letter, column_index_from_string
from df2 import extract_ghcn_daily

#Set up environment variable dictionary
EV = dict(os.environ)



#Need to write some screening functions to make sure the years necessary are
#available.


class stationDaily(object):
    """
    Object containing attributes and methods for analyzing GHCN .dly
    daily station data files.

    INPUTS
    stationID: ID code that identifies the .dly file
    NIPAdata: Tuple containing (phase, months, years)
    """

    def __init__(self, var, stationID, NIPAdata):
        self.stationID = stationID
        self.phase = NIPAdata[0]
        self.months = NIPAdata[1]
        self.years = NIPAdata[2]
        #self.n_years = self.years
        ###################################################################
        #_Organize data in a data frame from station data_#################
        fp = EV['GHCND_HCN'] + '/'+ stationID + '.dly'
        #extracts data
        all_data = extract_ghcn_daily(fp = fp)
        var_data = all_data[var]

        #reorders the dataset
        df = pd.DataFrame()
        for year in self.years:
            data = np.array([])
            for month in self.months:
                index = '%d-%d' % (year, month)
                data = np.concatenate((data, var_data[index].values))
            df['%d' % year] = data
        ###################################################################
        #10AM, def functions in def_init_ => threshold method, control bin size
        self.data = df
        self.Tdata = {}

    def threshold(self,thresholds):
        for threshold in thresholds:
            TDF = self.data.copy()
            #creates threshold data by turning off any data that lies beneath threshold
            idx = TDF <= threshold
            TDF[idx] = np.nan
            self.Tdata[str(threshold)] = TDF
            self.thresholds = thresholds
        return
    def otherstatistics(self):
        self.mean = {}
        for threshold in self.thresholds:
            t = {}
            for year in self.years:
                x = self.Tdata[str(threshold)]
                y = x[str(year)]
                t[str(year)] = y.mean()
            self.mean[str(threshold)] = t

        self.std = {}
        for threshold in self.thresholds:
            t = {}
            for year in self.years:
                x = self.Tdata[str(threshold)]
                y = x[str(year)]
                t[str(year)] = y.std()
            self.std[str(threshold)] = t

if __name__ == '__main__':
    def write_data(thresholds, stations):
        #write threshold data
        for threshold in thresholds:
            station_row = 3
            for station in stations:
                wb = Workbook()
                ws = wb.get_active_sheet() #create new worksheet
                ws.title = station.phase #name of the worksheet is the phase
                ws['A1'] = "Events over " + str(threshold) # events over threshold
                ws['A' + str(station_row)] = "Station " + station.stationID #write station name

                j=2 #column index start
                for i in station.years:
                    events = []
                    ws[get_column_letter(j) + '2'] = i #write the year to ws
                    tdata = station.Tdata[str(threshold)] #get data under the threshold key
                    yeardata = tdata[str(i)] #get the data for one year

                    for k in yeardata: #for each data entry, if it is not NaN add to list
                        if ~np.isnan(k):
                            events.append(k)

                    ws[get_column_letter(j) + '3'] = len(events)    #write number of events over threshold to ws

                    j = j + 1 # increment the column index
                    station_row = station_row + 1 #increment the row index
            wb.save('T' + str(threshold) + '.xlsx') #save the data according to the threshold value
        #write statistic data
        for threshold in thresholds:
            station_row = 3
            for station in stations:
                wb = Workbook()
                ws = wb.get_active_sheet() #create new worksheet
                ws.title = station.phase #name of the worksheet is the phase
                ws['A1'] = "Mean number of events over " + str(threshold) # events over threshold
                ws['A' + str(station_row)] = "Station " + station.stationID #write station name

                j=2 #column index start
                for i in station.years:
                    events = []
                    ws[get_column_letter(j) + '2'] = i #write the year to ws
                    mean = station.mean[str(threshold)][str(i)] #get data under the threshold key

                    ws[get_column_letter(j) + '3'] = mean    #write the mean for the station in that year

                    j = j + 1 # increment the column index
                    station_row = station_row + 1 #increment the row index
            wb.save('MEAN' + str(threshold) + '.xlsx') #save the data according to the threshold value
        #write STD
        for threshold in thresholds:
            station_row = 3
            for station in stations:
                wb = Workbook()
                ws = wb.get_active_sheet() #create new worksheet
                ws.title = station.phase #name of the worksheet is the phase
                ws['A1'] = "Standard Deviation for events over " + str(threshold) # events over threshold
                ws['A' + str(station_row)] = "Station " + station.stationID #write station name

                j=2 #column index start
                for i in station.years:
                    events = []
                    ws[get_column_letter(j) + '2'] = i #write the year to ws
                    STD = station.std[str(threshold)][str(i)] #get data under the threshold key

                    ws[get_column_letter(j) + '3'] = STD    #write the mean for the station in that year

                    j = j + 1 # increment the column index
                    station_row = station_row + 1 #increment the row index
            wb.save('STD' + str(threshold) + '.xlsx') #save the data according to the threshold value
        return
    ###_Station ID should come from station_info, produced currently by dw.py
    station_list = ['USC00410832']#,'USC00011084']
    thresholds = (0, 64) #threshold values
    NIPAdata = []
    NIPAdata.append(('El Nino', [3, 4, 5, 6],
                    [1926, 1931, 1941, 1942, 1958, 1966, 1973,
                    1983, 1992, 1995, 1998, 2010]))
    NIPAdata.append(('La Nina', [3, 4, 5, 6], [1925, 1934, 1943,
                    1950, 1951, 1955, 1956, 1957, 1971, 1972, 1974,
                    1975, 1976, 1989, 1999, 2000, 2008]))
    NIPAdata.append(('Neutral Positive', [3, 4, 5, 6], [1924, 1927,
                    1928, 1930, 1932, 1936, 1937, 1940, 1947, 1952, 1954,
                    1959, 1964, 1969, 1970, 1977, 1978, 1979, 1980, 1988,
                    1990, 1991, 1993, 1994, 2004, 2005, 2007]))
    NIPAdata.append(('Neutral Negative', [3, 4, 5, 6], [1921, 1922, 1923,
                        1929, 1933, 1935, 1938, 1939, 1944, 1945, 1946,
                        1948, 1949, 1953, 1960, 1961, 1962, 1963, 1965,
                        1967, 1968, 1981, 1982, 1984, 1985, 1986, 1996,
                        1997, 2001, 2002, 2006, 2009]))
    stations = []
    for stationID in station_list:
        for data in NIPAdata:
            var = 'PRCP'
            station = stationDaily(var, stationID, data)
            station.threshold(thresholds)
            station.otherstatistics()
        stations.append(station)
        print np.sum(np.isnan(station.data))
    #write_data(thresholds, stations)
